"""
migrate_to_supabase.py — Push Excel data to Supabase
Files:
  - 0. Web App [GSNB-CML] Master Data Dictionary (4).xlsx
  - 1. Web App [GSNB-BG40-2026] BOM Pricing (9).xlsx

Usage:
  py migrate_to_supabase.py [--dry-run] [--skip-master] [--skip-pricing]
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import os, re, json, argparse, datetime
import openpyxl
import requests

# ─── CONFIG ──────────────────────────────────────────────────────────────
MASTER_FILE  = r"c:\Users\pit008\Downloads\BOM\0. Web App [GSNB-CML] Master Data Dictionary (4).xlsx"
PRICING_FILE = r"c:\Users\pit008\Downloads\BOM\1. Web App [GSNB-BG40-2026] BOM Pricing (9).xlsx"

def load_env(path):
    env = {}
    with open(path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if '=' in line and not line.startswith('#'):
                k, v = line.split('=', 1)
                env[k.strip()] = v.strip()
    return env

ENV = load_env(r"c:\Users\pit008\Downloads\BOM-web\.env.local")
SUPABASE_URL     = ENV['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_ROLE_KEY = ENV['SUPABASE_SERVICE_ROLE_KEY']

BASE_HEADERS = {
    'apikey':        SERVICE_ROLE_KEY,
    'Authorization': f'Bearer {SERVICE_ROLE_KEY}',
    'Content-Type':  'application/json',
}

# ─── HELPERS ─────────────────────────────────────────────────────────────
def to_float(val):
    if val is None: return None
    try: return float(val)
    except (ValueError, TypeError):
        m = re.match(r'[\d.]+', str(val))
        return float(m.group()) if m else None

def to_str(val):
    return str(val).strip() if val is not None else None

def api_delete_all(table: str, pk: str = 'id', dry_run=False):
    """Delete all rows from a table."""
    if dry_run:
        print(f"    DELETE all from {table}")
        return
    url = f"{SUPABASE_URL}/rest/v1/{table}?{pk}=not.is.null"
    r = requests.delete(url, headers={**BASE_HEADERS, 'Prefer': 'return=minimal'})
    if r.status_code not in (200, 204):
        print(f"  [{table}] WARN delete: {r.status_code} {r.text[:100]}")

def api_insert(table: str, rows: list, dry_run=False, label=None):
    """Insert rows in batches of 500."""
    name = label or table
    if not rows:
        print(f"  [{name}] — 0 rows, skipping")
        return
    if dry_run:
        print(f"  [{name}] DRY-RUN — {len(rows)} rows to insert")
        print(f"    Sample: {rows[0]}")
        return

    batch_size = 500
    total = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        url = f"{SUPABASE_URL}/rest/v1/{table}"
        r = requests.post(url, headers={**BASE_HEADERS, 'Prefer': 'return=minimal'}, json=batch)
        if r.status_code not in (200, 201):
            print(f"  [{name}] ERROR {r.status_code}: {r.text[:300]}")
            return
        total += len(batch)
    print(f"  [{name}] ✓ inserted {total} rows")

def truncate_insert(table: str, rows: list, pk: str = 'id', dry_run=False):
    """Delete all then insert fresh rows."""
    api_delete_all(table, pk, dry_run)
    api_insert(table, rows, dry_run)

# ─── MASTER DATA ─────────────────────────────────────────────────────────
def migrate_master(wb, dry_run):
    print("\n=== MASTER DATA ===")

    # definition (code PK-ish, unique constraint — delete all first)
    ws = wb['Definition']
    rows = [{'code': to_str(r[2]), 'en_name': to_str(r[0]) or '', 'vn_name': to_str(r[1]) or ''}
            for r in list(ws.iter_rows(values_only=True))[1:] if r[2]]
    truncate_insert('definition', rows, dry_run=dry_run)

    # dm_quality, dm_color, dm_shape, dm_types, dm_category — (name, code)
    for sheet, table in [('DM_Quality','dm_quality'), ('DM_Color','dm_color'),
                          ('DM_Shape','dm_shape'), ('DM_Types','dm_types'), ('DM_Category','dm_category')]:
        ws = wb[sheet]
        seen_codes = set()
        rows = []
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if not r[0]: continue
            code = to_str(r[1]) or to_str(r[0])
            if not code or code in seen_codes: continue  # skip duplicates
            seen_codes.add(code)
            rows.append({'name': to_str(r[0]), 'code': code})
        truncate_insert(table, rows, dry_run=dry_run)

    # dm_size (actual columns: category, type, shape_code, color, quality, master_code,
    #          grade_id, pricing_unit, measurement_type, min_size, max_size, display_name,
    #          base_price, mk, diamond_price, vietnamese_name, full_name_en)
    ws = wb['1. DM_Size']
    rows = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[6]: continue  # grade_id required
        rows.append({
            'category':        to_str(r[0]),
            'type':            to_str(r[1]),
            'shape_code':      to_str(r[2]),
            'color':           to_str(r[3]),
            'quality':         to_str(r[4]),
            'master_code':     to_str(r[5]),
            'grade_id':        to_str(r[6]),
            'pricing_unit':    to_str(r[7]),
            'measurement_type': to_str(r[8]),
            'min_size':        to_float(r[9]),
            'max_size':        to_float(r[10]),
            'display_name':    to_str(r[11]),
            # col[12] = check_missing — no such column in DB, skip
            'base_price':      to_float(r[13]) or 0.0,
            'mk':              to_float(r[14]) or 0.0,      # mkup → mk
            'diamond_price':   to_float(r[15]),
            'vietnamese_name': to_str(r[16]),               # full_name_vi → vietnamese_name
            'full_name_en':    to_str(r[17]),
        })
    truncate_insert('dm_size', rows, pk='grade_id', dry_run=dry_run)

# ─── PRICING DATA ────────────────────────────────────────────────────────
def migrate_pricing(wb, dry_run):
    print("\n=== PRICING DATA ===")

    # gold_material (PK: price_date)
    ws = wb['Gold_Material']
    all_rows = list(ws.iter_rows(values_only=True))
    headers = all_rows[0]
    karat_map = {}
    for i, h in enumerate(headers):
        if h and str(h).startswith('Price_') and str(h).endswith('_gr'):
            label = str(h).replace('Price_', '').replace('_gr', '')
            karat_map[i] = label

    rows = []
    for r in all_rows[1:]:
        if r[0] is None: continue
        d = r[0]
        date_str = d.strftime('%Y-%m-%d') if isinstance(d, (datetime.datetime, datetime.date)) else str(d).strip()
        karat_prices = {karat_map[i]: float(r[i]) for i in karat_map if r[i] is not None}
        rows.append({
            'price_date':    date_str,
            'amark_gold_oz': to_float(r[1]),
            'amark_pt_oz':   to_float(r[2]),
            'amark_ag_oz':   to_float(r[3]),
            'loss_factor':   to_float(r[4]) or 1.06,
            'karat_prices':  json.dumps(karat_prices),
        })
    truncate_insert('gold_material', rows, pk='price_date', dry_run=dry_run)

    # stone_material (actual cols: group_code, grade_id, display_name, unit, type_input,
    #                  min_size, max_size, selling_price, base_price, mkup, full_name_vn, full_name_en)
    ws = wb['Stone_Material']
    rows = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[1]: continue  # grade_id required
        rows.append({
            'group_code':    to_str(r[0]),
            'grade_id':      to_str(r[1]),
            'display_name':  to_str(r[2]),
            'unit':          to_str(r[3]),
            'type_input':    to_str(r[4]),
            'min_size':      to_float(r[5]),
            'max_size':      to_float(r[6]),
            'selling_price': to_float(r[7]) or 0.0,
            'base_price':    to_float(r[8]) or 0.0,
            'mkup':          to_float(r[9]) or 0.0,
            'full_name_vn':  to_str(r[10]),              # full_name_vi → full_name_vn
            'full_name_en':  to_str(r[11]),
        })
    truncate_insert('stone_material', rows, pk='grade_id', dry_run=dry_run)

    # mk_price_list_type (PK: price_list_type)
    ws = wb['Price_List_Type_Re_MK_LC']
    rows = [{'price_list_type': to_str(r[0]), 'region': to_str(r[1]),
              'store': to_str(r[2]), 'logo_url': to_str(r[3])}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('mk_price_list_type', rows, pk='price_list_type', dry_run=dry_run)

    # mk_product_type (serial PK — actual cols: product_type, product_type_details_en, product_type_details_vi)
    ws = wb['Product_Type_MK_LC']
    rows = [{'product_type': to_str(r[0]), 'product_type_details_en': to_str(r[1]),
              'product_type_details_vi': to_str(r[2])}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('mk_product_type', rows, dry_run=dry_run)

    # mk_type_definition (serial PK — col: type_definition)
    ws = wb['Type_De_MK_LC']
    rows = [{'type_definition': to_str(r[1])}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[1]]
    truncate_insert('mk_type_definition', rows, dry_run=dry_run)

    # mk_color (serial PK — col: color)
    ws = wb['Color_MK_LC']
    rows = [{'color': to_str(r[2])}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[2]]
    truncate_insert('mk_color', rows, dry_run=dry_run)

    # mk_process_fee (serial PK — cols: unit_name, unit_price)
    ws = wb['Process_Fee_MK_LC']
    rows = [{'unit_name': to_str(r[0]), 'unit_price': to_float(r[1]) or 0.0}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('mk_process_fee', rows, dry_run=dry_run)

    # mk_cif_rate (PK: price_list_type)
    ws = wb['Price_List_Type_MK_LC']
    rows = [{'price_list_type': to_str(r[0]), 'cif_rate': to_float(r[1]) or 0.1}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('mk_cif_rate', rows, pk='price_list_type', dry_run=dry_run)

    # mk_store_markup (serial PK — actual col for prices: markups JSONB)
    ws = wb['Store_MK_LC']
    all_rows = list(ws.iter_rows(values_only=True))
    store_cols = [str(h).strip() for h in all_rows[0][2:] if h is not None]
    rows = []
    for r in all_rows[1:]:
        if r[0] is None: continue
        markups = {}
        for i, key in enumerate(store_cols):
            v = r[2 + i]
            if v is not None:
                markups[key] = float(v)
        rows.append({
            'value_from': to_float(r[0]) or 0.0,
            'value_to':   to_float(r[1]) or 0.0,
            'markups':    json.dumps(markups),             # prices → markups
        })
    truncate_insert('mk_store_markup', rows, dry_run=dry_run)

    # mk_price_gram (serial PK)
    ws = wb['Price_Gram_MK_LC']
    rows = [{'sp_type': to_str(r[0]), 'weight_from': to_float(r[1]) or 0.0,
              'weight_to': to_float(r[2]) or 0.0, 'markup_factor': to_float(r[3]) or 1.0,
              'additional_price': to_float(r[4]) or 0.0}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('mk_price_gram', rows, dry_run=dry_run)

    # salesperson (serial PK — actual col: salesperson_name)
    ws = wb['SalesPerson']
    rows = [{'salesperson_name': to_str(r[0]), 'email': to_str(r[1])}
             for r in list(ws.iter_rows(values_only=True))[1:] if r[0]]
    truncate_insert('salesperson', rows, dry_run=dry_run)

    # store (not stores! — actual: store_name, region)
    ws = wb['Stores']
    seen_stores = set()
    rows = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue
        name = to_str(r[0])
        if name in seen_stores: continue
        seen_stores.add(name)
        rows.append({'store_name': name, 'region': to_str(r[1])})
    truncate_insert('store', rows, dry_run=dry_run)

    # sys_config (PK: key — actual cols: key, value — no description column)
    ws = wb['SYS_CONFIG']
    target_keys = {'VND_RATE', 'PROCESS_FEE_LAPRAP', 'MANAGER_MAX_DISCOUNT'}
    seen = {}
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if r[0] is None: continue
        key = str(r[0]).strip()
        if key in target_keys and key not in seen:
            seen[key] = str(r[1]).strip() if r[1] is not None else '0'
    defaults = {'VND_RATE': '0', 'PROCESS_FEE_LAPRAP': '10', 'MANAGER_MAX_DISCOUNT': '20'}
    for k, v in defaults.items():
        if k not in seen:
            seen[k] = v
    rows = [{'key': k, 'value': v} for k, v in seen.items()]
    # sys_config uses text PK — upsert with merge-duplicates
    if not dry_run:
        url = f"{SUPABASE_URL}/rest/v1/sys_config"
        r = requests.post(url, headers={**BASE_HEADERS, 'Prefer': 'resolution=merge-duplicates'}, json=rows)
        if r.status_code not in (200, 201):
            print(f"  [sys_config] ERROR {r.status_code}: {r.text[:300]}")
        else:
            print(f"  [sys_config] ✓ upserted {len(rows)} rows")
    else:
        print(f"  [sys_config] DRY-RUN — {len(rows)} rows: {rows}")

# ─── BOM DATA ────────────────────────────────────────────────────────────
def fmt_dt(val):
    """Convert Excel datetime/date to ISO string."""
    if val is None: return None
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, datetime.date):
        return val.isoformat()
    return str(val).strip() or None

def fmt_date(val):
    """Convert Excel date to DATE string (YYYY-MM-DD)."""
    if val is None: return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime('%Y-%m-%d')
    return str(val).strip()[:10] or None


def migrate_bom(wb, dry_run):
    print("\n=== BOM DATA ===")

    # ── bom (from BOM_DB) ─────────────────────────────────────────────
    ws = wb['BOM_DB']
    bom_rows = []
    valid_ids = set()
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue  # skip blank rows
        bom_id = str(r[0]).strip()
        valid_ids.add(bom_id)
        bom_rows.append({
            'bom_id':         bom_id,
            'timestamp':      fmt_dt(r[1]),
            'date':           fmt_date(r[2]),
            'product_type':   to_str(r[3]),
            'so_mo':          to_str(r[4]),
            'model':          to_str(r[5]),
            'total_stone_qty': int(r[6]) if r[6] is not None else 0,
            'total_stone_ctw': to_float(r[7]) or 0.0,
            'labor_hours':    to_float(r[8]) or 0.0,
            'price_list_type': to_str(r[9]),
            'sp_type':        to_str(r[10]),
            'cost_gold':      to_float(r[11]) or 0.0,
            'cost_stones':    to_float(r[12]) or 0.0,
            'cost_labor':     to_float(r[13]) or 0.0,
            'cost_subtotal':  to_float(r[14]) or 0.0,
            'cost_cif':       to_float(r[15]) or 0.0,
            'cost_total':     to_float(r[16]) or 0.0,
            'sell_price':     to_float(r[17]) or 0.0,
            'note':           to_str(r[18]),
            'img1':           to_str(r[19]),
            'img2':           to_str(r[20]),
            'img3':           to_str(r[21]),
            'folder_url':     to_str(r[22]),
            'created_by':     to_str(r[23]) or 'imported',
            'updated_at':     fmt_dt(r[24]),
            'updated_by':     to_str(r[25]),
            'customer_name':  to_str(r[26]),
            'discount_pct':   to_float(r[27]) or 0.0,
            'discount_price': to_float(r[28]) or 0.0,
            'sales_person':   to_str(r[29]),
            'store':          to_str(r[30]),
        })

    # Delete existing bom (cascade deletes bom_gold + bom_stone via FK)
    api_delete_all('bom', pk='bom_id', dry_run=dry_run)
    api_insert('bom', bom_rows, dry_run=dry_run)

    # ── bom_gold (from BOM_Gold) ──────────────────────────────────────
    ws = wb['BOM_Gold']
    gold_rows = []
    skipped_gold = 0
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]: continue
        bom_id = str(r[0]).strip()
        if bom_id not in valid_ids:
            skipped_gold += 1
            continue
        gold_rows.append({
            'bom_id':    bom_id,
            'idx':       int(r[1]) if r[1] is not None else 1,
            'gold_type': to_str(r[2]),   # strip trailing spaces
            'color':     to_str(r[3]),
            'weight':    to_float(r[4]) or 0.0,
        })
    if skipped_gold:
        print(f"  [bom_gold] Skipped {skipped_gold} rows with no matching BOM_ID")
    api_insert('bom_gold', gold_rows, dry_run=dry_run)

    # ── bom_stone (from BOM_Stone) ────────────────────────────────────
    ws = wb['BOM_Stone']
    stone_rows = []
    skipped_stone = 0
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[1]: continue
        bom_id = str(r[1]).strip()
        if bom_id not in valid_ids:
            skipped_stone += 1
            continue
        grade = to_str(r[4])
        input_type = to_str(r[9])
        # Skip pure placeholder rows (no group, grade is '-')
        if not to_str(r[3]) and grade in ('-', None):
            skipped_stone += 1
            continue
        # Normalize invalid input_type '-' → None
        if input_type not in ('mm', 'ct', None):
            input_type = None
        stone_rows.append({
            'bom_id':     bom_id,
            'idx':        int(r[2]) if r[2] is not None else 1,
            'group_code': to_str(r[3]),
            'grade_id':   grade,
            'size':       to_float(r[5]),
            'ctw1pc':     to_float(r[6]) or 0.0,
            'qty':        int(r[7]) if r[7] is not None else 0,
            'tl_hot':     to_float(r[8]) or 0.0,
            'input_type': input_type,
            'gia_ban':    to_float(r[10]) or 0.0,
        })
    if skipped_stone:
        print(f"  [bom_stone] Skipped {skipped_stone} rows with no matching BOM_ID")
    api_insert('bom_stone', stone_rows, dry_run=dry_run)


# ─── MAIN ────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run',      action='store_true', help='Preview without writing')
    parser.add_argument('--skip-master',  action='store_true')
    parser.add_argument('--skip-pricing', action='store_true')
    parser.add_argument('--skip-bom',     action='store_true')
    parser.add_argument('--bom-only',     action='store_true', help='Only migrate BOM tables')
    args = parser.parse_args()

    dry = args.dry_run
    if dry:
        print("DRY-RUN mode — no data will be written\n")

    if not args.skip_master and not args.bom_only:
        print(f"Loading: {MASTER_FILE}")
        wb1 = openpyxl.load_workbook(MASTER_FILE, data_only=True)
        migrate_master(wb1, dry)

    if not args.skip_pricing and not args.bom_only:
        print(f"\nLoading: {PRICING_FILE}")
        wb2 = openpyxl.load_workbook(PRICING_FILE, data_only=True)
        migrate_pricing(wb2, dry)

    if not args.skip_bom:
        print(f"\nLoading BOM data: {PRICING_FILE}")
        wb_bom = openpyxl.load_workbook(PRICING_FILE, data_only=True)
        migrate_bom(wb_bom, dry)

    print("\n✓ Migration complete")

if __name__ == '__main__':
    main()
